import openpyxl, os

# Using chdir() to change active working directory
os.chdir(r'C:\Users\mwj\Desktop')

# Using load_workbook() method to load a workbook from working directory
workbook = openpyxl.load_workbook('example.xlsx')

# Using get_sheet_names() method to list sheets in workbook
# workbook.get_sheet_names()

# Using get_sheet_by_name() to create sheet object
sheet = workbook['Sheet1'] # .get_sheet_by_name('Sheet1')

# Using sheet['A1'] syntax to access a cell
# and value method to access its value
print(sheet['C1'].value)

# Using cell method to access cell
# and value method to access its value
print(sheet.cell(row=1, column=3).value)
