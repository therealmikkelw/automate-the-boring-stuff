# Creates a new Excel workbook with an amortization table based
# on the arguments hovedstol, indbetaling per termin and årlig rente.
#
# Termin every 3 months is assumed.

import openpyxl, logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# add an argument filename='myProgramLog.txt' to logging.basicConfig to save debug logs to a file
logging.disable(logging.CRITICAL) # this line disables debug logging

logging.debug('Start of program')

def amortTable( hovedstol, indbetaling, aarligRente ):

    logging.debug('Start of amortTable( %d, %d, %f )' % (hovedstol, indbetaling, aarligRente))
    # Create a new workbook and sheet object
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet.title = 'My amortization table'
    logging.debug('Workbook created')
    
    # Create first row holding 'saldo', 'indbetaling', 'renter', 'afdrag', 'slut saldo'
    startSaldo = -1 * hovedstol
    sheet['B2'] = startSaldo
    sheet['C2'] = indbetaling
    sheet['D2'] = round(startSaldo * aarligRente / 4)
    sheet['E2'] = round(indbetaling + (startSaldo * aarligRente / 4))
    slutSaldo = round(startSaldo + indbetaling + (startSaldo * aarligRente / 4))
    sheet['F2'] = slutSaldo
    logging.debug('First row created')
    logging.debug("'slut saldo' is %d" % (sheet['F2'].value))

    # Repeat until 'slut saldo' is 0
    i = 3
    while slutSaldo < 0:
        startSaldo = slutSaldo
        sheet.cell(row=i, column=2).value = startSaldo
        sheet.cell(row=i, column=3).value = indbetaling
        sheet.cell(row=i, column=4).value = round(startSaldo * aarligRente / 4)
        sheet.cell(row=i, column=5).value = round(indbetaling + (startSaldo * aarligRente / 4))
        slutSaldo = round(startSaldo + indbetaling + (startSaldo * aarligRente / 4))
        sheet.cell(row=i, column=6).value = slutSaldo
        logging.debug('Row %d created, slutSaldo is %d' % (i, slutSaldo))
        i += 1

    logging.debug('Data table done')
    # Create header and 'termin'
    sheet['A1'] = 'Termin'
    sheet['B1'] = 'Start saldo'
    sheet['C1'] = 'Indbetaling'
    sheet['D1'] = 'Renter'
    sheet['E1'] = 'Afdrag'
    sheet['F1'] = 'Slut saldo'
    logging.debug('Table headers created')

    for j in range(2, i):
        sheet.cell(row=j, column=1).value = 'Termin ' + str(j - 1)
    logging.debug("'Termin' column done")    

    # Save the workbook
    wb.save('myAmortTable.xlsx')
    logging.debug('Workbook is saved')  

amortTable( hovedstol=442749, indbetaling=21075, aarligRente=0.05 )
