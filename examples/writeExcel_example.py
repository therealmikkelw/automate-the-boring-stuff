import openpyxl, os

# Using Workbook() method to create a new workbook object
wb = openpyxl.Workbook()

# List sheet names
# wb.get_sheet_names()

# Access sheet object
sheet = wb['Sheet'] # .get_sheet_by_name('Sheet')

# Assign values to cells
sheet['A1'] = 42
sheet['A2'] = 'Hello'

# Using create_sheet() to create a new sheet
sheet2 = wb.create_sheet(index=0, title='My other sheet')

# Rename the sheet using title method
sheet2.title = 'My new sheet name'

# Using save() method to store the workbook to harddrive
os.chdir(r'C:\Users\mwj\Documents\Automate the boring stuff')
wb.save('example.xlsx')
